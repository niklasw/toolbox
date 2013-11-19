#!/usr/bin/env python
#
# Example script converting ascii data to xlsx file
# Indata is a set of files (from commandline e.g. *.data)
# each on the form
# [data name] [separator] [value] like
#
# compressor mfap = 1853.0074462890625
# filter_inlet mfap = 3502.183349609375
# filter_outlet mfap = 2665.8974609375
# inlet mfap = 4439.9638671875
# outlet mfap = 1485.9603271484375
# porous_inner mfap = 2704.845458984375
# porous_outer mfap = 3140.668212890625


import os,sys,re
from openpyxl.workbook import Workbook

def Error(s,sig=1):
    print '\nError %s!\n' % s
    sys.exit(sig)

def Warn(s):
    output = 'Warning %s!' % s
    print '\n'+'='*len(output)
    print output
    print '='*len(output)

def Info(s):
    print '\t%s' % s

class indata(dict):
    def __init__(self,files=[],separator='=',comments='#%'):
        dict.__init__(self)
        self.keys = files
        self.separator = separator
        self.comments = comments
        self.readAllFiles()

    def readFile(self,fileName):
        with open(fileName) as fh:
            lines = [ a for a in fh.readlines() if not a.strip()[0] in self.comments ]
        return dict([a.split(self.separator) for a in lines])

    def readAllFiles(self):
        for key in self.keys:
            self[key] = self.readFile(key)

class sheetNav:
    def __init__(self):
        alphabet = map(chr, range(65, 91))
        doubleAlphabet = list(alphabet)
        for c in alphabet:
            doubleAlphabet += [ c+a for a in alphabet ]

        self.alphabet = doubleAlphabet
        self.charMap = dict( zip(doubleAlphabet, range(len(doubleAlphabet))) )
        self.intMap  = dict( zip(map(str, range(1,1001)),range(0,1000)) )

    def charRangeToIntList(self,charRange, mapHash):
        columns = set()
        ranges = charRange.split(',')
        for r in ranges:
            if r in doubleAlphabet:
                columns.add(mapHash[r])
            else:
                begin,end = r.split(':')
                columns = columns.union(sorted(mapHash.values())[mapHash[begin]:mapHash[end]+1])
        return list(sorted(columns))

    def rowRange(self,charRange):
        return charRangeToIntList(charRange,self.charMap)

    def columnRange(self,charRange):
        return charRangeToIntList(charRange,self.intMap)

    def cell(self,rownum,colnum):
        return self.alphabet[colnum-1]+str(rownum)

def getArgs():
    from optparse import OptionParser
    descString = """
    Python thing to convert ascii data files to xlsx
    """

    parser=OptionParser(description=descString)
    parser.add_option('-o','--output',dest='output',default='output.xlsx',help='Output xlsx file')
    parser.add_option('-s','--separator',dest='separator',default='=',help='Column separator')
    parser.add_option('-c','--commentchars',dest='commentchars',default='#%',help='String of chars leading comment lines')
    #parser.add_option('-S','--sheets',dest='sheets',action='store_true',default=False,help='Have each input file on different sheet')

    (opt,arg)=parser.parse_args()
    validArgs = list()

    def argError(s):
        s = '* ERROR: %s. *' % s
        n=len(s)
        print '\n\t%s\n\t%s\n\t%s\n' % (n*'*',s,n*'*')
        parser.print_help()
        sys.exit(1)

    def validateOption(option, test, msg='Invalid argument', allowed=[]):
        try:    option = test(option)
        except: argError('%s; got %s' % (msg,option))
        if allowed and not option in allowed:
            argError('%s; got %s. Allowed values are %s' % (msg,option,allowed))
        return option

    validArgs = [ a for a in arg if os.path.isfile(a) ]

    return opt,validArgs


wb = Workbook()
wb.remove_sheet(wb.get_active_sheet())

opt,fileNames = getArgs()

dataDict = indata(files=fileNames, separator=opt.separator, comments=opt.commentchars)

nav = sheetNav()

sheet = wb.create_sheet()
sheet.title = 'Reports data'

startCol = 1
lastKeys = []
for dataset, data in dataDict.items():
    row = 1
    keys = data.keys()
    nextCol = startCol
    if lastKeys != data.keys():
        Info('New data set in file %s ' % dataset)
        nextCol += 1
        for i,row in enumerate(data.keys()):
            sheet.cell(nav.cell(i+2,startCol)).value = row
        startCol += 1

    sheet.cell(nav.cell(1,nextCol)).value = str(dataset)
    for i,value in enumerate(data.values()):
        thisCell = sheet.cell(nav.cell(i+2,nextCol))
        thisCell.value = float(value)
    startCol += 1
    lastKeys = data.keys()

wb.save(opt.output)
