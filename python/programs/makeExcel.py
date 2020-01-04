#!/usr/bin/env python3
# -*- coding: utf-8 -*-


class sheetNav:
    def __init__(self):
        alphabet = map(chr, range(65, 91))
        doubleAlphabet = list(alphabet)
        for c in alphabet:
            doubleAlphabet += [ c+a for a in alphabet ]

        self.alphabet = doubleAlphabet
        self.charMap = dict( zip(doubleAlphabet, range(len(doubleAlphabet))) )
        self.intMap  = dict( zip(map(str, range(1,1001)),range(0,1000)) )

    def charRangeToIntList(self,charRange, mapHash):
        columns = set()
        ranges = charRange.split(',')
        for r in ranges:
            if r in doubleAlphabet:
                columns.add(mapHash[r])
            else:
                begin,end = r.split(':')
                columns = columns.union(sorted(mapHash.values())[mapHash[begin]:mapHash[end]+1])
        return list(sorted(columns))

    def rowRange(self,charRange):
        return charRangeToIntList(charRange,self.charMap)

    def columnRange(self,charRange):
        return charRangeToIntList(charRange,self.intMap)

    def cell(self,rownum,colnum):
        return self.alphabet[colnum-1]+str(rownum)


if __name__ == '__main__':

    from openpyxl.workbook import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    nav = sheetNav()
    sheet = wb.create_sheet()
    sheet.title = 'Tunnels'
    print(nav.cell(2,2))

    #sheet.cell(nav.cell(2,2)).value = "Hejsan!"
    sheet.cell(2,2).value = "Hejsan!"
    sheet.cell(2,3).value = "Picture link"
    sheet.cell(2,3).hyperlink = 'hpics\pic.png'

    wb.save('/home/niklas/F/makeExcel.xlsx')



    

